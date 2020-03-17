# -*- coding: utf-8 -*-
"""
Created on Thu Nov  7 01:51:25 2019

@author: Antonio
"""

import os
import csv
from openpyxl import workbook #pip install openpyxl
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font, Alignment
import shutil
from datetime import datetime,date
import random
import string
import psycopg2
import credentials

query = '''with cdi_futuro as 
(SELECT date,ctos_negociados,
SUM(contratos_negociados) FILTER (WHERE mercadoria = 'DI1') as "DI1",
SUM(contratos_negociados) FILTER (WHERE mercadoria = 'D11') as "D11",
SUM(contratos_negociados) FILTER (WHERE mercadoria = 'D12') as "D12",
SUM(contratos_negociados) FILTER (WHERE mercadoria = 'D13') as "D13",
SUM(contratos_negociados) FILTER (WHERE mercadoria = 'D14') as "D14",
SUM(contratos_negociados) FILTER (WHERE mercadoria = 'IDI') as "IDI",
SUM(contratos_negociados) FILTER (WHERE mercadoria = 'OC1') as "OC1",
SUM(contratos_negociados) FILTER (WHERE mercadoria = 'DAP') as "DAP",
SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'DI1') as "DI1a",
SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'D11') as "D11a",
SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'D12') as "D12a",
SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'D13') as "D13a",
SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'D14') as "D14a",
SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'IDI') as "IDIa",
SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'OC1') as "OC1a",
SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'DAP') as "DAPa"
FROM
(SELECT format('%s-%s-01',ano,mes)::date date, round(SUM(contratos_negociados) over (partition by ano,mes)/1000) ctos_negociados ,
mercadoria, contratos_negociados, contratos_em_aberto
FROM
(SELECT date_part('year', data_pregao) ano, date_part('month', data_pregao) mes, 
mercadoria, sum(contratos_negociados) contratos_negociados, sum(contratos_em_aberto) contratos_em_aberto,produto
	FROM bmf.vw_result where produto = 'CDI FUTURO'
	GROUP BY 1,2,3,produto) a) b GROUP BY date,ctos_negociados),
---
cdi_curto_longo as
(SELECT date_part('year', data_pregao) ano, date_part('month', data_pregao) mes,
produto, sum(contratos_negociados) FILTER (WHERE dias_saque_ate_vencimento < 290) contratos_negociados_curto, 
sum(contratos_negociados) FILTER (WHERE dias_saque_ate_vencimento >= 290) contratos_negociados_longo,
sum(contratos_em_aberto) FILTER (WHERE dias_saque_ate_vencimento < 290) contratos_em_aberto_curto,
sum(contratos_em_aberto) FILTER (WHERE dias_saque_ate_vencimento >= 290) contratos_em_aberto_longo
	FROM bmf.vw_result where produto = 'CDI FUTURO'
	group by 1,2,3),
---
dolar_futuro as
(SELECT date,ctos_negociados,
SUM(contratos_negociados) FILTER (WHERE mercadoria = 'DOL' and mercado = 2) as "DOL",
SUM(contratos_negociados) FILTER (WHERE mercadoria = 'DR1') as "DR1",
SUM(contratos_negociados) FILTER (WHERE mercadoria = 'FRP') as "FRP",
SUM(contratos_negociados) FILTER (WHERE mercadoria = 'EUR') as "EUR",
SUM(contratos_negociados) FILTER (WHERE mercadoria = 'DOL' and mercado = 3) as "OPC",
SUM(contratos_negociados) FILTER (WHERE mercadoria = 'WDL') as "WDL",
SUM(contratos_negociados) FILTER (WHERE mercadoria = 'WDO') as "WDO",
SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'DOL' and mercado = 2) as "DOLa",
SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'DR1') as "DR1a",
SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'FRP') as "FRPa",
SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'EUR') as "EURa",
SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'DOL' and mercado = 3) as "OPCa",
SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'WDL') as "WDLa",
SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'WDO') as "WDOa"
FROM
(SELECT format('%s-%s-01',ano,mes)::date date, round(SUM(contratos_negociados) over (partition by ano,mes)/1000) ctos_negociados ,
mercadoria,mercado, contratos_negociados, contratos_em_aberto
FROM
(SELECT date_part('year', data_pregao) ano, date_part('month', data_pregao) mes, 
mercadoria,mercado, sum(contratos_negociados) contratos_negociados, sum(contratos_em_aberto) contratos_em_aberto,produto
	FROM bmf.vw_result where produto = 'DOLAR FUTURO'
	group by 1,2,3,4,produto) a) b GROUP BY date,ctos_negociados),
---
ibov_futuro as
(SELECT date,ctos_negociados,
SUM(contratos_negociados) FILTER (WHERE mercadoria = 'IND') as "IND",
SUM(contratos_negociados) FILTER (WHERE mercadoria = 'IR1') as "IR1",
SUM(contratos_negociados) FILTER (WHERE mercadoria = 'BRI') as "BRI",
SUM(contratos_negociados) FILTER (WHERE mercadoria = 'ISP') as "ISP",
SUM(contratos_negociados) FILTER (WHERE mercadoria = 'WIN') as "WIN",
SUM(contratos_negociados) FILTER (WHERE mercadoria = 'WD1') as "WD1",
SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'IND') as "INDa",
SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'IR1') as "IR1a",
SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'BRI') as "BRIa",
SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'ISP') as "ISPa",
SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'WIN') as "WINa",
SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'WD1') as "WD1a"
FROM
(SELECT format('%s-%s-01',ano,mes)::date date, round(SUM(contratos_negociados) over (partition by ano,mes)/1000) ctos_negociados ,
mercadoria, contratos_negociados, contratos_em_aberto
FROM
(SELECT date_part('year', data_pregao) ano, date_part('month', data_pregao) mes, 
mercadoria, sum(contratos_negociados) contratos_negociados, sum(contratos_em_aberto) contratos_em_aberto,produto
	FROM bmf.vw_result where produto = 'IBOV FUTURO'
	group by 1,2,3,produto) a) b GROUP BY date,ctos_negociados),
---
cupom_cambial as 
(SELECT date,ctos_negociados,
SUM(contratos_negociados) FILTER (WHERE mercadoria = 'DDI') as "DDI",
SUM(contratos_negociados) FILTER (WHERE mercadoria = 'FRC') as "FRC",
SUM(contratos_negociados) FILTER (WHERE mercadoria = 'SCC') as "SCC",
SUM(contratos_negociados) FILTER (WHERE mercadoria = 'SCS') as "SCS",
SUM(contratos_negociados) FILTER (WHERE mercadoria = 'DCO') as "DCO",
SUM(contratos_negociados) FILTER (WHERE mercadoria = 'FRO') as "FRO",
SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'DDI') as "DDIa",
SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'FRC') as "FRCa",
SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'SCC') as "SCCa",
SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'SCS') as "SCSa",
SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'DCO') as "DCOa",
SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'FRO') as "FROa"
FROM
(SELECT format('%s-%s-01',ano,mes)::date date, round(SUM(contratos_negociados) over (partition by ano,mes)/1000) ctos_negociados ,
mercadoria, contratos_negociados, contratos_em_aberto
FROM
(SELECT date_part('year', data_pregao) ano, date_part('month', data_pregao) mes, 
mercadoria, sum(contratos_negociados) contratos_negociados, sum(contratos_em_aberto) contratos_em_aberto
	,produto
	FROM bmf.vw_result where produto = 'CUPOM CAMBIAL FUTURO'
	group by 1,2,3,produto) a) b GROUP BY date,ctos_negociados),
du as
(SELECT count(distinct data_pregao),date_part('year', data_pregao) ano, date_part('month', data_pregao) mes FROM bmf.vw_result WHERE extract('ISODOW' FROM data_pregao) < 6
	GROUP BY 2,3)

---------------
SELECT cdi_futuro.date, cdi_futuro.ctos_negociados,  du.count du, null,
"DI1","D11","D12","D13","D14","IDI","OC1","DAP",null,"DI1a","D11a","D12a","D13a","D14a","IDIa","OC1a","DAPa",null,
contratos_negociados_curto,contratos_negociados_longo,null,contratos_em_aberto_curto,contratos_em_aberto_longo,null,
"DOL","DR1","FRP","EUR","OPC","WDL","WDO",null,"DOLa","DR1a","FRPa","EURa","OPCa","WDLa","WDOa",null,
"IND","IR1","BRI","ISP","WIN","WD1",null,"INDa","IR1a","BRIa","ISPa","WINa","WD1a",null,
"DDI","FRC","SCC","SCS","DCO","FRO",null,"DDIa","FRCa","SCCa","SCSa","DCOa","FROa"
FROM cdi_futuro 
JOIN cdi_curto_longo ON date_part('year', cdi_futuro.date) = cdi_curto_longo.ano AND date_part('month', cdi_futuro.date) = cdi_curto_longo.mes
JOIN dolar_futuro ON cdi_futuro.date = dolar_futuro.date
JOIN ibov_futuro ON cdi_futuro.date = ibov_futuro.date
JOIN cupom_cambial ON cdi_futuro.date = cupom_cambial.date 
JOIN du ON date_part('year', cdi_futuro.date) = du.ano AND date_part('month', cdi_futuro.date) = du.mes
ORDER BY 1'''

DATABASE, HOST, USER, PASSWORD = credentials.setDatabaseLogin()

### conecta no banco de dados
db_conn = psycopg2.connect("dbname='{}' user='{}' host='{}' password='{}'".format(DATABASE, USER, HOST, PASSWORD))
cursor = db_conn.cursor()
print('Connected to the database')

file = 'model_derivativos.xlsx'
outdir = '/home/ubuntu/scripts/load-dados-bmf/'

cursor.execute(query)
result = [item for item in cursor.fetchall()]

wb = load_workbook(filename=outdir+file, read_only=False)
sheet = wb['2019-2020']
last_row = sheet.max_row + 1
col = 2

for i in range(len(result)):
    for j in range(len(result[i])):
        sheet.cell(row = last_row+i, column = col+j).value = result[i][j]
        sheet.cell(row = last_row+i, column = col+j).font = Font(size=10)
        sheet.cell(row = last_row+i, column = col+j).alignment = Alignment(horizontal='center')

wb.save(outdir+'DadosDerivativos.xlsx')
wb.close()

# def writeExcel(file,outdir):
    # def findLastRow(csvfile,sheet):
        # last_row = sheet.max_row
        # if csvfile == 'Aquisição de veículos.csv':
            # col = 1
        # elif csvfile == 'Cheque especial.csv':
            # col = 6
        # elif csvfile == 'Crédito pessoal consignado público.csv':
            # col = 11
        # elif csvfile == 'Crédito pessoal consignado INSS.csv':
            # col = 21
        # elif csvfile == 'Crédito pessoal consignado privado.csv':
            # col = 26
        # elif csvfile == 'Crédito pessoal não-consignado.csv':
            # col = 36
        # elif csvfile == 'Cartão de crédito - rotativo total.csv':
            # col = 56
        # elif csvfile == 'Cartão de crédito - rotativo em curso normal.csv':
            # col = 66
        # elif csvfile == 'Cartão de crédito - rotativo em atraso.csv':
            # col = 71
        # elif csvfile == 'Cartão de crédito - parcelado.csv':
            # col = 76
        # else:
            # print(csvfile)
            # raise
        # while sheet.cell(column=col, row=last_row).value is None and last_row > 0:
            # last_row -= 1
        # return col,last_row

    # csvfiles = [f for f in os.listdir(outdir) if f.endswith('.csv')]
    # wb = load_workbook(filename=outdir+file, read_only=False, keep_vba=True)
    # sheet = wb['2019-2020']
    # stylename = ''.join([random.choice(string.ascii_letters + string.digits) for i in range(10)])
    # nsmmmyy=NamedStyle(name=stylename, number_format="DD/MMM/YY")
    # nsmmmyy.font = Font(size=10)
    # nsmmmyy.alignment = Alignment(horizontal='center')
    # for csvfile in csvfiles:
        # with open(outdir+csvfile, "r",newline="\n", encoding="utf-8") as ifile:
            # reader = csv.reader(ifile, delimiter=',')
            # header = next(reader, None)
            # col,last_row = findLastRow(csvfile,sheet)
            # j = 0
            # for line in reader:
                # for i in range(4):
                    # if i == 0:
                        # line[i] = datetime.strptime(line[i], '%d/%m/%Y').date()
                        # sheet.cell(row = last_row+j, column = col+i).value = line[i]
                        # sheet.cell(row = last_row+j, column = col+i).style = nsmmmyy
                    # elif i == 2 or i == 3:
                        # line[i] = float(line[i].replace('.','').replace(',','.'))
                        # sheet.cell(row = last_row+j, column = col+i).value = line[i]
                        # sheet.cell(row = last_row+j, column = col+i).font = Font(size=10)
                    # else:
                        # sheet.cell(row = last_row+j, column = col+i).value = line[i]
                        # sheet.cell(row = last_row+j, column = col+i).font = Font(size=10)
                # j += 1
    # wb.save(outdir+'new_'+file)
    # wb.close()
    # for csvfile in csvfiles:
        # os.remove(outdir+csvfile)
    # os.remove(outdir+file)
    # shutil.move(outdir+'new_'+file, outdir+file)
