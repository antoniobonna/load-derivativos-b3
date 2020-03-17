# -*- coding: utf-8 -*-
"""
Created on Thu Nov  7 01:51:25 2019

@author: Antonio
"""

import os
import csv
from openpyxl import workbook #pip install openpyxl
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font, Alignment, Color
import shutil
from datetime import datetime,date
import random
import string
import psycopg2
import credentials

query_di = '''SELECT date_part('month',date) mes, date_part('year',date) ano,date,
null,null,null,null,null,null,null,null,null,null,null,
du.count du,
null,null,null,null,null,null,null,null,null,null,null,null,null,null,null,
coalesce(SUM(contratos_negociados) FILTER (WHERE mercadoria = 'DI1'),0) as "DI1",
coalesce(SUM(contratos_negociados) FILTER (WHERE mercadoria = 'D11'),0) as "D11",
coalesce(SUM(contratos_negociados) FILTER (WHERE mercadoria = 'D12'),0) as "D12",
coalesce(SUM(contratos_negociados) FILTER (WHERE mercadoria = 'D13'),0) as "D13",
coalesce(SUM(contratos_negociados) FILTER (WHERE mercadoria = 'D14'),0) as "D14",
coalesce(SUM(contratos_negociados) FILTER (WHERE mercadoria = 'IDI'),0) as "IDI",
coalesce(SUM(contratos_negociados) FILTER (WHERE mercadoria = 'OC1'),0) as "OC1",
coalesce(SUM(contratos_negociados) FILTER (WHERE mercadoria = 'DAP'),0) as "DAP",
null,
coalesce(SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'DI1'),0) as "DI1a",
coalesce(SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'D11'),0) as "D11a",
coalesce(SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'D12'),0) as "D12a",
coalesce(SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'D13'),0) as "D13a",
coalesce(SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'D14'),0) as "D14a",
coalesce(SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'IDI'),0) as "IDIa",
coalesce(SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'OC1'),0) as "OC1a",
coalesce(SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'DAP'),0) as "DAPa",
null,
contratos_negociados_curto/1000,contratos_negociados_longo/1000,null,contratos_em_aberto_curto/1000,contratos_em_aberto_longo/1000
FROM
(SELECT format('%s-%s-01',ano,mes)::date date, round(SUM(contratos_negociados) over (partition by ano,mes)/1000) ctos_negociados ,
mercadoria, contratos_negociados, contratos_em_aberto
FROM
(SELECT date_part('year', data_pregao) ano, date_part('month', data_pregao) mes,
mercadoria, sum(contratos_negociados) contratos_negociados, sum(contratos_em_aberto) contratos_em_aberto,produto
	FROM bmf.vw_result where produto = 'CDI FUTURO' AND data_pregao >= '2019-10-01'
	GROUP BY 1,2,3,produto) a) b 
	JOIN (SELECT count(distinct data_pregao),date_part('year', data_pregao) ano, date_part('month', data_pregao) mes FROM bmf.vw_result
	GROUP BY 2,3) du ON du.ano = date_part('year',date) AND du.mes = date_part('month',date)
	JOIN (SELECT date_part('year', data_pregao) ano, date_part('month', data_pregao) mes,
produto, sum(contratos_negociados) FILTER (WHERE dias_saque_ate_vencimento < 290) contratos_negociados_curto, 
sum(contratos_negociados) FILTER (WHERE dias_saque_ate_vencimento >= 290) contratos_negociados_longo,
sum(contratos_em_aberto) FILTER (WHERE dias_saque_ate_vencimento < 290) contratos_em_aberto_curto,
sum(contratos_em_aberto) FILTER (WHERE dias_saque_ate_vencimento >= 290) contratos_em_aberto_longo
	FROM bmf.vw_result where produto = 'CDI FUTURO' AND data_pregao >= '2019-10-01'
	group by 1,2,3) cdi_curto_longo ON cdi_curto_longo.ano = date_part('year',date) AND cdi_curto_longo.mes = date_part('month',date)
	GROUP BY 1,2,3,ctos_negociados,du,contratos_negociados_curto,contratos_negociados_longo,contratos_em_aberto_curto,contratos_em_aberto_longo
	order by date'''

query_dol = '''SELECT date_part('month',date) mes, date_part('year',date) ano,date,
null,null,null,null,
du.count du,null,null,ctos_negociados,
null,null,null,null,null,null,null,null,
coalesce(SUM(contratos_negociados) FILTER (WHERE mercadoria = 'DOL' and mercado = 2),0) as "DOL",
coalesce(SUM(contratos_negociados) FILTER (WHERE mercadoria = 'DR1'),0) as "DR1",
coalesce(SUM(contratos_negociados) FILTER (WHERE mercadoria = 'FRP'),0) as "FRP",
coalesce(SUM(contratos_negociados) FILTER (WHERE mercadoria = 'EUR'),0) as "EUR",
coalesce(SUM(contratos_negociados) FILTER (WHERE mercadoria = 'DOL' and mercado = 3),0) as "OPC",
coalesce(SUM(contratos_negociados) FILTER (WHERE mercadoria = 'WDL'),0) as "WDL",
coalesce(SUM(contratos_negociados) FILTER (WHERE mercadoria = 'WDO'),0) as "WDO",
null,
round(coalesce(SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'DOL' and mercado = 2),0)/du.count) as "DOLa",
round(coalesce(SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'DR1'),0)/du.count) as "DR1a",
round(coalesce(SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'FRP'),0)/du.count) as "FRPa",
round(coalesce(SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'EUR'),0)/du.count) as "EURa",
round(coalesce(SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'DOL' and mercado = 3),0)/du.count) as "OPCa",
round(coalesce(SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'WDL'),0)/du.count) as "WDLa",
round(coalesce(SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'WDO'),0)/du.count) as "WDOa"
FROM
(SELECT format('%s-%s-01',ano,mes)::date date, round(SUM(contratos_negociados) over (partition by ano,mes)) ctos_negociados ,
mercadoria,mercado, contratos_negociados, contratos_em_aberto
FROM
(SELECT date_part('year', data_pregao) ano, date_part('month', data_pregao) mes, 
mercadoria,mercado, sum(contratos_negociados) contratos_negociados, sum(contratos_em_aberto) contratos_em_aberto,produto
	FROM bmf.vw_result where produto = 'DOLAR FUTURO' AND data_pregao >= '2019-10-01'
	group by 1,2,3,4,produto) a) b 
	JOIN (SELECT count(distinct data_pregao),date_part('year', data_pregao) ano, date_part('month', data_pregao) mes FROM bmf.vw_result
	GROUP BY 2,3) du ON du.ano = date_part('year',date) AND du.mes = date_part('month',date)
	GROUP BY date,ctos_negociados,du
	ORDER BY date'''

query_cupom = '''SELECT date_part('year',date) ano,date,
null,null,null,null,null,
du.count du,
null,null,null,null,null,null,null,null,null,null,null,
coalesce(SUM(contratos_negociados) FILTER (WHERE mercadoria = 'DDI'),0) as "DDI",
coalesce(SUM(contratos_negociados) FILTER (WHERE mercadoria = 'FRC'),0) as "FRC",
coalesce(SUM(contratos_negociados) FILTER (WHERE mercadoria = 'SCC'),0) as "SCC",
coalesce(SUM(contratos_negociados) FILTER (WHERE mercadoria = 'SCS'),0) as "SCS",
coalesce(SUM(contratos_negociados) FILTER (WHERE mercadoria = 'DCO'),0) as "DCO",
coalesce(SUM(contratos_negociados) FILTER (WHERE mercadoria = 'FRO'),0) as "FRO",
null,
coalesce(SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'DDI'),0) as "DDIa",
coalesce(SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'FRC'),0) as "FRCa",
coalesce(SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'SCC'),0) as "SCCa",
coalesce(SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'SCS'),0) as "SCSa",
coalesce(SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'DCO'),0) as "DCOa",
coalesce(SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'FRO'),0) as "FROa"
FROM
(SELECT format('%s-%s-01',ano,mes)::date date, round(SUM(contratos_negociados) over (partition by ano,mes)/1000) ctos_negociados ,
mercadoria, contratos_negociados, contratos_em_aberto
FROM
(SELECT date_part('year', data_pregao) ano, date_part('month', data_pregao) mes, 
mercadoria, sum(contratos_negociados) contratos_negociados, sum(contratos_em_aberto) contratos_em_aberto,produto
	FROM bmf.vw_result where produto = 'CUPOM CAMBIAL FUTURO' AND data_pregao >= '2019-10-01'
	group by 1,2,3,produto) a) b 
	JOIN (SELECT count(distinct data_pregao),date_part('year', data_pregao) ano, date_part('month', data_pregao) mes FROM bmf.vw_result
	GROUP BY 2,3) du ON du.ano = date_part('year',date) AND du.mes = date_part('month',date)
	GROUP BY date,ctos_negociados,du
	ORDER BY date'''

query_ibov = '''SELECT date_part('year',date) ano,date,
null,null,null,null,
du.count du,
null,null,null,null,null,null,null,null,null,null,null,null,null,
coalesce(SUM(contratos_negociados) FILTER (WHERE mercadoria = 'IND'),0) as "IND",
coalesce(SUM(contratos_negociados) FILTER (WHERE mercadoria = 'IR1'),0) as "IR1",
coalesce(SUM(contratos_negociados) FILTER (WHERE mercadoria = 'BRI'),0) as "BRI",
coalesce(SUM(contratos_negociados) FILTER (WHERE mercadoria = 'ISP'),0) as "ISP",
coalesce(SUM(contratos_negociados) FILTER (WHERE mercadoria = 'WIN'),0) as "WIN",
coalesce(SUM(contratos_negociados) FILTER (WHERE mercadoria = 'WD1'),0) as "WD1",
null,
coalesce(SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'IND'),0) as "INDa",
coalesce(SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'IR1'),0) as "IR1a",
coalesce(SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'BRI'),0) as "BRIa",
coalesce(SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'ISP'),0) as "ISPa",
coalesce(SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'WIN'),0) as "WINa",
coalesce(SUM(contratos_em_aberto) FILTER (WHERE mercadoria = 'WD1'),0) as "WD1a"
FROM
(SELECT format('%s-%s-01',ano,mes)::date date, round(SUM(contratos_negociados) over (partition by ano,mes)/1000) ctos_negociados ,
mercadoria, contratos_negociados, contratos_em_aberto
FROM
(SELECT date_part('year', data_pregao) ano, date_part('month', data_pregao) mes, 
mercadoria, sum(contratos_negociados) contratos_negociados, sum(contratos_em_aberto) contratos_em_aberto,produto
	FROM bmf.vw_result where produto = 'IBOV FUTURO' AND data_pregao >= '2019-10-01'
	group by 1,2,3,produto) a) b 
	JOIN (SELECT count(distinct data_pregao),date_part('year', data_pregao) ano, date_part('month', data_pregao) mes FROM bmf.vw_result
	GROUP BY 2,3) du ON du.ano = date_part('year',date) AND du.mes = date_part('month',date)
	GROUP BY date,ctos_negociados,du
	ORDER BY date'''

DATABASE, HOST, USER, PASSWORD = credentials.setDatabaseLogin()

### conecta no banco de dados
db_conn = psycopg2.connect("dbname='{}' user='{}' host='{}' password='{}'".format(DATABASE, USER, HOST, PASSWORD))
cursor = db_conn.cursor()
print('Connected to the database')

file = 'BMF_SOTP.xlsx'
indir = '/home/ubuntu/scripts/load-dados-bmf/'

wb = load_workbook(filename=indir+file, read_only=False)

sheet = wb['DI_FUT']
last_row = 160#sheet.max_row + 1
col = 76

cursor.execute(query_di)
result = [item for item in cursor.fetchall()]

for i in range(len(result)):
    for j in range(len(result[i])):
        if result[i][j]:
            sheet.cell(row = last_row+i, column = col+j).value = result[i][j]
            sheet.cell(row = last_row+i, column = col+j).font = Font(size=10)
            sheet.cell(row = last_row+i, column = col+j).alignment = Alignment(horizontal='center')
            if (col+j) >= 106:
                sheet.cell(row = last_row+i, column = col+j).font = Font(color="0000FF")

sheet = wb['DOL_FUT']
last_row = 159#sheet.max_row + 1
col = 50

cursor.execute(query_dol)
result = [item for item in cursor.fetchall()]

for i in range(len(result)):
    for j in range(len(result[i])):
        if result[i][j]:
            sheet.cell(row = last_row+i, column = col+j).value = result[i][j]
            sheet.cell(row = last_row+i, column = col+j).font = Font(size=10)
            sheet.cell(row = last_row+i, column = col+j).alignment = Alignment(horizontal='center')
            if (col+j) >= 69:
                sheet.cell(row = last_row+i, column = col+j).font = Font(color="0000FF")

sheet = wb['CUPOM_FX']
last_row = 159#sheet.max_row + 1
col = 54

cursor.execute(query_cupom)
result = [item for item in cursor.fetchall()]

for i in range(len(result)):
    for j in range(len(result[i])):
        if result[i][j]:
            sheet.cell(row = last_row+i, column = col+j).value = result[i][j]
            sheet.cell(row = last_row+i, column = col+j).font = Font(size=10)
            sheet.cell(row = last_row+i, column = col+j).alignment = Alignment(horizontal='center')


sheet = wb['IBOV']
last_row = 159#sheet.max_row + 1
col = 58

cursor.execute(query_ibov)
result = [item for item in cursor.fetchall()]

for i in range(len(result)):
    for j in range(len(result[i])):
        if result[i][j]:
            sheet.cell(row = last_row+i, column = col+j).value = result[i][j]
            sheet.cell(row = last_row+i, column = col+j).font = Font(size=10)
            sheet.cell(row = last_row+i, column = col+j).alignment = Alignment(horizontal='center')
            if (col+j) >= 78:
                sheet.cell(row = last_row+i, column = col+j).font = Font(color="0000FF")

wb.save(indir+'new_BMF_SOTP.xlsx')
wb.close()

# def writeExcel(file,indir):
    # def findLastRow(csvfile,sheet):
        # last_row = sheet.max_row
        # if csvfile == 'Aquisição de veículos.csv':
            # col = 1
        # elif csvfile == 'Cheque especial.csv':
            # col = 6
        # elif csvfile == 'Crédito pessoal consignado público.csv':
            # col = 11
        # elif csvfile == 'Crédito pessoal consignado INSS.csv':
            # col = 21
        # elif csvfile == 'Crédito pessoal consignado privado.csv':
            # col = 26
        # elif csvfile == 'Crédito pessoal não-consignado.csv':
            # col = 36
        # elif csvfile == 'Cartão de crédito - rotativo total.csv':
            # col = 56
        # elif csvfile == 'Cartão de crédito - rotativo em curso normal.csv':
            # col = 66
        # elif csvfile == 'Cartão de crédito - rotativo em atraso.csv':
            # col = 71
        # elif csvfile == 'Cartão de crédito - parcelado.csv':
            # col = 76
        # else:
            # print(csvfile)
            # raise
        # while sheet.cell(column=col, row=last_row).value is None and last_row > 0:
            # last_row -= 1
        # return col,last_row

    # csvfiles = [f for f in os.listdir(indir) if f.endswith('.csv')]
    # wb = load_workbook(filename=indir+file, read_only=False, keep_vba=True)
    # sheet = wb['2019-2020']
    # stylename = ''.join([random.choice(string.ascii_letters + string.digits) for i in range(10)])
    # nsmmmyy=NamedStyle(name=stylename, number_format="DD/MMM/YY")
    # nsmmmyy.font = Font(size=10)
    # nsmmmyy.alignment = Alignment(horizontal='center')
    # for csvfile in csvfiles:
        # with open(indir+csvfile, "r",newline="\n", encoding="utf-8") as ifile:
            # reader = csv.reader(ifile, delimiter=',')
            # header = next(reader, None)
            # col,last_row = findLastRow(csvfile,sheet)
            # j = 0
            # for line in reader:
                # for i in range(4):
                    # if i == 0:
                        # line[i] = datetime.strptime(line[i], '%d/%m/%Y').date()
                        # sheet.cell(row = last_row+j, column = col+i).value = line[i]
                        # sheet.cell(row = last_row+j, column = col+i).style = nsmmmyy
                    # elif i == 2 or i == 3:
                        # line[i] = float(line[i].replace('.','').replace(',','.'))
                        # sheet.cell(row = last_row+j, column = col+i).value = line[i]
                        # sheet.cell(row = last_row+j, column = col+i).font = Font(size=10)
                    # else:
                        # sheet.cell(row = last_row+j, column = col+i).value = line[i]
                        # sheet.cell(row = last_row+j, column = col+i).font = Font(size=10)
                # j += 1
    # wb.save(indir+'new_'+file)
    # wb.close()
    # for csvfile in csvfiles:
        # os.remove(indir+csvfile)
    # os.remove(indir+file)
    # shutil.move(indir+'new_'+file, indir+file)
